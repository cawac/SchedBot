import json
import os
import re
from datetime import datetime, date, timedelta
import openpyxl

def is_yellow(cell):
    fill = cell.fill
    color = fill.fgColor

    if color.type == "rgb":
        return color.rgb.upper() == "FFFFFF00"
    return False

def enrich_with(func_to_run):
    def decorator(func):
        def wrapper(*args, **kwargs):
            result = func(*args, **kwargs)
            enriched = {
                **result,
                **func_to_run(*args, **kwargs)
            }
            return enriched
        return wrapper
    return decorator


def get_date_from_sheet(sheet, row, column):
    while sheet.cell(row, 3).value is None:
        row -= 1
    return {"date": sheet.cell(row, 3).value}


def replacer(match):
    content = match.group(1)
    if content == "L" or content.startswith("pr"):
        return f"({content})"
    return content


# @enrich_with(get_date_from_sheet)
@enrich_with(lambda sheet, row, column: {"groups": [sheet.cell(row, 5).value]})
@enrich_with(lambda sheet, row, column: {"lesson_number": int(sheet.cell(2, column).value)})
def get_lesson_info(sheet, row, column):
    lesson_info: str = sheet.cell(row, column).value
    if isinstance(lesson_info, str):
        lesson_info = lesson_info.strip().replace("\n", "")

    auditorium_match = re.search(r"aud\s*(\d+)", lesson_info, re.IGNORECASE)
    auditorium = auditorium_match.group(1) if auditorium_match else None

    lesson_info = re.sub(r"aud\s*\d+", "", lesson_info, flags=re.IGNORECASE).strip()
    lesson_info = lesson_info.split('(')[0].strip()
    return {
        "subject_name": lesson_info,
        "lesson_type": "Lecture" if "(L)" in lesson_info else "Practice",
        "auditorium": auditorium,
    }


class Parser:
    def __init__(self):
        self.__file_path: str | None = None
        self.__sheet = None
        self.__amount_of_groups: int | None = None
        self.working_area: dict | None = None
        self.__ignore_words = ["holiday"]

    def load_schedule(self, schedule_file_path: str, **working_area):
        self.__file_path = schedule_file_path
        self.working_area = working_area
        self.pre_process()
        self.get_groups_amount()

    def pre_process(self):
        wb = openpyxl.load_workbook(self.__file_path)
        self.__sheet = wb.active
        self.expand_merged_cells()
        self.correct_cells()
        self.remove_redundant()

    def expand_merged_cells(self):
        merged_ranges = list(self.__sheet.merged_cells.ranges)

        for merged_range in merged_ranges:
            min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
            top_left_value = self.__sheet.cell(row=min_row, column=min_col).value

            self.__sheet.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)

            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    self.__sheet.cell(row=row, column=col).value = top_left_value

    def correct_cells(self):
        for row in range(self.working_area.get("min_row", 1), self.working_area.get("max_row", 500) + 1):
            for column in range(self.working_area.get("min_column", 1), self.working_area.get("max_column", 15) + 1):
                cell = self.__sheet.cell(row=row, column=column).value
                if cell is None:
                    continue
                self.__sheet.cell(row=row, column=column).value = re.sub(r"\(([^()]+)\)", replacer, cell)

    def parse_lessons_for_day(self, date: datetime | date) -> list:
        cell_coords = self.get_row_coordinates_by_day(date)
        lessons = []

        if cell_coords is None:
            return lessons

        date_row = cell_coords

        for row in range(date_row, date_row + self.__amount_of_groups):
            for column in range(self.working_area.get("min_column", 6), self.working_area.get("max_column", 12) + 1):
                cell = self.__sheet.cell(row=row, column=column)
                if self.__sheet.cell(row, column).value is not None or is_yellow(cell):
                    lessons.append(get_lesson_info(self.__sheet, row, column))

        cleaned_lessons = None

        for i in range(len(lessons)):
            if cleaned_lessons is None:
                cleaned_lessons = [lessons[i]]
            else:
                for j in range(len(cleaned_lessons)):
                    if (  # cleaned_lessons[j]['date'],
                        cleaned_lessons[j]["lesson_number"],
                        cleaned_lessons[j]["subject_name"]) == (
                            # lessons[i]['date'],
                            lessons[i]["lesson_number"],
                            lessons[i]["subject_name"]):
                        cleaned_lessons[j]["groups"].extend(lessons[i]['groups'])
                        break
                else:
                    cleaned_lessons.append(lessons[i])
        return cleaned_lessons

    def get_row_coordinates_by_day(self, target_day: datetime) -> int | None:
        if isinstance(target_day, datetime):
            target_day = target_day.date()
        elif not isinstance(target_day, date):
            return None

        for row in self.__sheet.iter_rows():
            for cell in row:
                cell_value = cell.value

                if isinstance(cell_value, datetime):
                    cell_date = cell_value.date()
                elif isinstance(cell_value, date):
                    cell_date = cell_value
                else:
                    continue

                if cell_date == target_day:
                    return cell.row

        return None

    def get_groups_amount(self):
        monday_row = self.get_row_coordinates_by_day(datetime(2025, 3, 3))
        tuesday_row = self.get_row_coordinates_by_day(datetime(2025, 3, 4))
        self.__amount_of_groups = tuesday_row - monday_row if monday_row and tuesday_row else 0

    def remove_redundant(self):
        for row in range(self.working_area.get("min_row", 1), self.working_area.get("max_row", 500) + 1):
            for column in range(self.working_area.get("min_column", 1), self.working_area.get("max_column", 15) + 1):
                cell = self.__sheet.cell(row=row, column=column)
                cell_value: str = str(cell.value) if cell.value else None

                if cell_value is None:
                    continue

                if any(ignored_word in cell_value.lower() for ignored_word in self.__ignore_words):
                    cell.value = None
                    continue

                cell.value = re.sub(r"\(([^()]+)\)", replacer, cell_value)

    def get_all_groups(self) -> list:
        result: list[str] = []
        for i in range(4, 4 + self.__amount_of_groups):
            result.append(self.__sheet.cell(row=i, column=5).value)

        return result


def get_all_file_paths(directory):
    file_paths = []
    for root, dirs, files in os.walk(directory):
        for filename in files:
            full_path = os.path.join(root, filename)
            file_paths.append(full_path)
    return file_paths


# example script
# if __name__ == "__main__":
#     parser = Parser()
#     for file_path in get_all_file_paths("schedules"):
#         parser.load_schedule(file_path, min_column=6, min_row=4, max_column=13, max_row=863)
#         data = parser.parse_lessons_for_day(datetime(2025, 3, 11) + timedelta(days=0))
