from dataclasses import dataclass
import os
from datetime import datetime

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

import xlwings as xw

def replace_formulas_with_excel(path: str, save_as: str = None):
    app = xw.App(visible=False)
    wb = app.books.open(path)
    wb.api.Application.CalculateFullRebuild()  # полный пересчёт формул
    if save_as:
        wb.save(save_as)
    else:
        wb.save()
    wb.close()
    app.quit()

DATE_COLUMN = 0
GROUP_COLUMN = 0
LESSON_NUMBER_ROW = 0
FIRST_WEEK_ROW = 0

def get_cell_with_text(worksheet: Worksheet, text: str, up: int, left: int, down: int, right: int):
    text = text.lower()

    for row in worksheet.iter_rows(up, down, left, right):
        for cell in row:
            if cell.value is None:
                continue

            value_str = str(cell.value).lower().strip()

            if text == value_str:
                return cell

    return None

def find_anchors(worksheet: Worksheet):
    global GROUP_COLUMN, DATE_COLUMN, LESSON_NUMBER_ROW, FIRST_WEEK_ROW
    FIRST_WEEK_ROW = get_cell_with_text(worksheet, '1 week', 1, 1, 10, 10).row
    group_cell = get_cell_with_text(worksheet, 'group', 1, 1, 10, 10)
    GROUP_COLUMN = group_cell.column
    DATE_COLUMN = get_cell_with_text(worksheet, 'vilnius time', 1, 1, 10, 10).column
    LESSON_NUMBER_ROW = group_cell.row

# replace with strategy for merged cells
def expand_merged_cells(worksheet: Worksheet):
    for merged_range in tuple(worksheet.merged_cells.ranges):
        left, up, right, down = merged_range.bounds
        value = worksheet.cell(row=up, column=left).value
        worksheet.unmerge_cells(start_row=up, start_column=left, end_row=down, end_column=right)

        for row in worksheet.iter_rows(up, down, left, right):
            for cell in row:
                cell.value = value

def format_merged_cells(worksheet: Worksheet):
    for merged_range in tuple(worksheet.merged_cells.ranges):
        left, up, right, down = merged_range.bounds
        worksheet.unmerge_cells(start_row=up, start_column=left, end_row=down, end_column=right)
        worksheet.cell(up, left, format_cell(worksheet, worksheet.cell(up, left)))
        worksheet.merge_cells(start_row=up, start_column=left, end_row=down, end_column=right)

def format_worksheet(worksheet: Worksheet):
    for row in worksheet.iter_rows():
        for cell in row:
            format_cell(worksheet, cell)

@dataclass
class Rule:
    rule: callable
    count_function: callable


rules = (
    Rule(lambda x: x.replace('\n', ' '), lambda x: x.count('\n')),
    Rule(lambda x: x.replace('\t', ' '), lambda x: x.count('\t')),
    Rule(lambda x: x.strip(), lambda x: 0),
)

def format_cell(worksheet: Worksheet, cell):
    cell_value = cell.value
    if cell_value is None:
        return

    if isinstance(cell_value, str):
        count = cell_value.count('\n') + cell_value.count('\t') + cell_value.count('  ') + cell_value.count('\xa0')
        if count > 0:
            print(f'formated: \\n = {cell_value.count('\n')}, \\t = {cell_value.count('\t')}, "  " = {cell_value.count('  ')}, "\\xa0 = {cell_value.count('\xa0')}"')

        cell_value = cell_value.replace('\n', ' ')
        cell_value = cell_value.replace('\t', ' ')
        cell_value = cell_value.replace('\xa0', ' ')
        cell_value = cell_value.strip()
        while '  ' in cell_value:
            cell_value = cell_value.replace('  ', ' ')

    cell.value = cell_value

def replace_formulas_with_values(filename: str, output: str = None):
    wb_formulas = openpyxl.load_workbook(filename, data_only=False)
    wb_values = openpyxl.load_workbook(filename, data_only=True)

    for sheet_name in wb_formulas.sheetnames:
        ws_formulas = wb_formulas[sheet_name]
        ws_values = wb_values[sheet_name]

        for row in ws_formulas.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    cell.value = ws_values[cell.coordinate].value

    if not output:
        output = filename
    wb_formulas.save(output)

class Formater:
    def __init__(self) -> None:
        self.__ignore_words: tuple = ('holiday',)
        self.__ignore_rules: list = []  # add rules to ignore cell with yellow color
        self.__ignore_worksheet: tuple = ('Staff load',)
        self.__file_path: str | None = None
        self.__workbook: Workbook | None = None
        self.__current_worksheet: Worksheet | None = None

    def load_file(self, file_path: str):
        self.__file_path = file_path
        self.__workbook: Workbook = openpyxl.load_workbook(self.__file_path)
        replace_formulas_with_values(file_path, file_path)

    def format_file(self, func):
        for worksheet in self.__workbook.worksheets:
            if worksheet.title in self.__ignore_worksheet:
                continue

            self.__current_worksheet = worksheet
            func(self.__current_worksheet)
            self.__workbook.save(self.__file_path)

def get_all_file_paths(directory):
    file_paths = []
    for root, dirs, files in os.walk(directory):
        for filename in files:
            full_path = os.path.join(root, filename)
            file_paths.append(full_path)
    return file_paths


if __name__ == '__main__':
    formater = Formater()

    for file_path in get_all_file_paths('schedules'):
        formater.load_file(file_path)
        # formater.format_file(format_merged_cells)
        formater.format_file(format_worksheet)
