import os
import re
from time import sleep
from typing import Optional, Dict

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


DATE_COLUMN = 0
GROUP_COLUMN = 0
LESSON_NUMBER_ROW = 0
FIRST_WEEK_ROW = 0

def is_yellow(cell):
    fill = cell.fill
    color = fill.fgColor

    if color.type == 'rgb':
        return color.rgb.upper() == 'FFFFFF00'
    return False

def enrich_with(func_to_run):
    def decorator(func):
        def wrapper(*args, **kwargs):
            result = func(*args, **kwargs)
            enriched = {
                **result,
                **func_to_run(*args, **kwargs)
            }
            return enriched
        return wrapper
    return decorator


def get_date_from_sheet(worksheet: Worksheet, row: int, column: int):
    while worksheet.cell(row, DATE_COLUMN).value is None:
        row -= 1
    return {'date': worksheet.cell(row, DATE_COLUMN).value}


def replacer(match):
    content = match.group(1)
    if content == 'L' or content.startswith('pr'):
        return f'({content})'
    return content

def get_groups(worksheet: Worksheet, row: int, column: int):
    return {'groups': [worksheet.cell(row, GROUP_COLUMN).value]}

def get_lesson_number(worksheet: Worksheet, row: int, column: int):
    if worksheet.cell(LESSON_NUMBER_ROW, column).value is int:
        return {'lesson_number': int(worksheet.cell(LESSON_NUMBER_ROW, column).value)}
    else:
        return {'lesson_number': None}


# move clearing part into formater
def clear_string(text: str) -> str:
    text = re.sub(r'\s+', ' ', text).lower()
    return text.strip()

# cell_content = subject type auditorium
# date
# group/s
# number of lesson
# subject
# type of lesson
# auditorium?
# online?
@enrich_with(get_date_from_sheet)
@enrich_with(get_groups)  # refine get_groups
@enrich_with(get_lesson_number)
def get_lesson_info(worksheet: Worksheet, row: int, column: int) -> Optional[Dict]:
    lesson_info: str = worksheet.cell(row, column).value
    if not lesson_info:
        return {}

    lesson_info = clear_string(lesson_info)
    is_subject_name = True
    subject_name = []
    lesson_type = ''
    extra_data = []
    for word in lesson_info.split(' '):
        if word.startswith('pr') and len(word) <= 3:
            lesson_type = 'practice'
            is_subject_name = False
            continue

        if word.startswith('lc') and len(word) <= 3:
            lesson_type = 'lecture'
            is_subject_name = False
            continue

        if is_subject_name:
            subject_name.append(word)
        else:
            extra_data.append(word)

    subject_name = ' '.join(subject_name)
    extra_data = ' '.join(extra_data)
    return {
        'subject_name': subject_name,
        'lesson_type': lesson_type,
        'extra_data': extra_data,
    }

def get_cell_with_text(worksheet: Worksheet, text: str, up: int, left: int, down: int, right: int):
    text = text.lower()

    for row in worksheet.iter_rows(up, down, left, right):
        for cell in row:
            if cell.value is None:
                continue

            value_str = str(cell.value).lower().strip()

            if text == value_str:
                return cell

    return None

def merge_dicts(dictionary1: dict[str, ...], dictionary2: dict[str, ...]) -> None:
    for key, value in dictionary2.items():
        if key in dictionary1:
            dictionary1[key].append(value)
        else:
            dictionary1[key] = value


def find_anchors(worksheet: Worksheet):
    global GROUP_COLUMN, DATE_COLUMN, LESSON_NUMBER_ROW, FIRST_WEEK_ROW
    FIRST_WEEK_ROW = get_cell_with_text(worksheet, '1 week', 1, 1, 10, 10).row
    group_cell = get_cell_with_text(worksheet, 'group', 1, 1, 10, 10)
    GROUP_COLUMN = group_cell.column
    DATE_COLUMN = get_cell_with_text(worksheet, 'vilnius time', 1, 1, 10, 10).column
    LESSON_NUMBER_ROW = group_cell.row

# replace with strategy for merged cells
def expand_merged_cells(worksheet: Worksheet):
    for merged_range in tuple(worksheet.merged_cells.ranges):
        left, up, right, down = merged_range.bounds
        value = worksheet.cell(row=up, column=left).value
        worksheet.unmerge_cells(start_row=up, start_column=left, end_row=down, end_column=right)

        for row in worksheet.iter_rows(up, down, left, right):
            for cell in row:
                cell.value = value


def get_groups(worksheet: Worksheet):
    first_week_row = FIRST_WEEK_ROW
    group_column = GROUP_COLUMN
    groups = [str(worksheet.cell(first_week_row, group_column).value)]
    first_week_row += 1

    while str(worksheet.cell(first_week_row, group_column).value) not in groups:
        groups.append(str(worksheet.cell(first_week_row, group_column).value))
        first_week_row += 1

    return groups


class ScheduleParser:
    def __init__(self) -> None:
        self.__ignore_words: tuple = ('holiday',)
        self.__ignore_rules: list = []  # add rules to ignore cell with yellow color
        self.__ignore_worksheet: tuple = ('Staff load',)

    def load_file(self, file_path: str):
        workbook: Workbook = openpyxl.load_workbook(file_path)
        return self.load_workbook(workbook)

    def load_workbook(self, workbook: Workbook):
        data: dict[str, ...] = {}
        for worksheet in workbook.worksheets:
            if worksheet.title in self.__ignore_worksheet:
                continue

            merge_dicts(data, self.load_worksheet(worksheet))

    def load_worksheet(self, worksheet: Worksheet) -> Dict[str, ...]:
        expand_merged_cells(worksheet)
        find_anchors(worksheet)
        result = {}
        result['groups'] = get_groups(worksheet)
        result['lessons'] = self.get_all_lessons(worksheet)
        return result

    def get_all_lessons(self, worksheet: Worksheet):
        result = []
        for row in range(FIRST_WEEK_ROW, worksheet.max_row + 1):
            for col in range(GROUP_COLUMN + 1, worksheet.max_column + 1):
                if worksheet.cell(row, col).value is not None:
                    result.append(get_lesson_info(worksheet, row, col))
                    print(get_lesson_info(worksheet, row, col))
                    sleep(1)

        return result

def get_all_file_paths(directory):
    file_paths = []
    for root, dirs, files in os.walk(directory):
        for filename in files:
            full_path = os.path.join(root, filename)
            file_paths.append(full_path)
    return file_paths


if __name__ == '__main__':
    parser = ScheduleParser()

    for file_path in get_all_file_paths('schedules'):
        result = parser.load_file(file_path)